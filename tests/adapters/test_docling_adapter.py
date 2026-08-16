import pytest
from openpyxl import Workbook
from contextlib import asynccontextmanager

from lib.adapters.docling import (
    ConversionStatus,
    DoclingAdapter,
    SPREADSHEET_MARKDOWN_MARKER,
    _chat_completions_model,
    _chat_completions_url,
)


async def _conversion_results(adapter, files):
    return [result async for result in adapter.extract_documents(files)]


def test_docling_sdk_import_available():
    from docling.document_converter import DocumentConverter

    assert DocumentConverter is not None


def test_rtf_conversion_extracts_text_without_docling(tmp_path):
    path = tmp_path / "readme.rtf"
    path.write_text(
        r"{\rtf1\ansi\ansicpg1252 This folder contains:\par Confidential caf\'e9.}",
        encoding="latin-1",
    )

    markdown = DoclingAdapter._convert_rtf_sync(str(path))

    assert markdown == "This folder contains:\nConfidential café.\n"


@pytest.mark.asyncio
async def test_rtf_processing_bypasses_docling_converter(monkeypatch, tmp_path):
    path = tmp_path / "readme.rtf"
    path.write_text(r"{\rtf1\ansi Plain text.}", encoding="latin-1")

    @asynccontextmanager
    async def slot(*_args, **_kwargs):
        yield None

    monkeypatch.setattr("lib.services_gateway.gateway.slot", slot)
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_sync",
        staticmethod(lambda _path: pytest.fail("Docling converter should not handle RTF")),
    )

    text = await DoclingAdapter()._process_single_file(str(path), "readme.rtf")

    assert text == "Plain text.\n"


@pytest.mark.asyncio
async def test_extract_documents_ignores_nonempty_files_without_text(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "image-only.pdf"
    path.write_bytes(b"%PDF image-only")

    async def no_text(*_args, **_kwargs):
        return ""

    monkeypatch.setattr(DoclingAdapter, "_process_single_file", no_text)

    results = await _conversion_results(
        DoclingAdapter(),
        [{"filename": path.name, "local_path": path}],
    )

    assert results[0].status is ConversionStatus.IGNORED_EMPTY
    assert results[0].reason == "no_extractable_text"


@pytest.mark.asyncio
async def test_extract_documents_keeps_conversion_exceptions_fatal(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "broken.pdf"
    path.write_bytes(b"%PDF broken")

    async def fail(*_args, **_kwargs):
        raise RuntimeError("converter unavailable")

    monkeypatch.setattr(DoclingAdapter, "_process_single_file", fail)

    results = await _conversion_results(
        DoclingAdapter(),
        [{"filename": path.name, "local_path": path}],
    )

    assert results[0].status is ConversionStatus.FAILED
    assert "converter unavailable" in results[0].error


@pytest.mark.asyncio
async def test_extract_documents_ignores_unsupported_formats(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "logo.eps"
    path.write_bytes(b"%!PS-Adobe EPSF")

    async def unsupported(*_args, **_kwargs):
        raise RuntimeError("File format not allowed: logo.eps")

    monkeypatch.setattr(DoclingAdapter, "_process_single_file", unsupported)

    results = await _conversion_results(
        DoclingAdapter(),
        [{"filename": path.name, "local_path": path}],
    )

    assert results[0].status is ConversionStatus.IGNORED_EMPTY
    assert results[0].reason == "unsupported_format"


@pytest.mark.asyncio
async def test_extract_documents_skips_known_unsupported_extensions(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "logo.eps"
    path.write_bytes(b"%!PS-Adobe EPSF")

    async def unexpected_convert(*_args, **_kwargs):
        raise AssertionError("unsupported extension should be skipped")

    monkeypatch.setattr(DoclingAdapter, "_process_single_file", unexpected_convert)

    results = await _conversion_results(
        DoclingAdapter(),
        [{"filename": path.name, "local_path": path}],
    )

    assert results[0].status is ConversionStatus.IGNORED_EMPTY
    assert results[0].reason == "unsupported_format"


def test_spreadsheet_conversion_omits_formatting_only_cells(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "financial model"
    sheet["A1"] = "Revenue"
    sheet["B2"] = 1200
    sheet.merge_cells("A5:XFD5")

    path = tmp_path / "model.xlsx"
    workbook.save(path)

    markdown = DoclingAdapter._convert_spreadsheet_sync(str(path))

    assert markdown.startswith(SPREADSHEET_MARKDOWN_MARKER)
    assert "## financial model" in markdown
    assert "Revenue" in markdown
    assert "1200" in markdown
    assert len(markdown) < 1_000
    assert "XFD" not in markdown


def test_spreadsheet_conversion_omits_formulas_without_cached_values(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Amount"
    sheet["A2"] = "Seed"
    sheet["B2"] = "=40+60"

    path = tmp_path / "normal.xlsx"
    workbook.save(path)

    markdown = DoclingAdapter._convert_spreadsheet_sync(str(path))

    assert "Seed" in markdown
    assert "=40+60" not in markdown


@pytest.mark.asyncio
@pytest.mark.parametrize("extension", [".xls", ".xlsx", ".xlsm"])
async def test_spreadsheet_processing_bypasses_docling_converter(monkeypatch, tmp_path, extension):
    path = tmp_path / f"model{extension}"
    path.write_bytes(b"workbook")

    @asynccontextmanager
    async def slot(*_args, **_kwargs):
        yield None

    monkeypatch.setattr("lib.services_gateway.gateway.slot", slot)
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_spreadsheet_sync",
        staticmethod(lambda _path: "compact values\n"),
    )
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_sync",
        staticmethod(lambda _path: pytest.fail("Docling converter should not handle spreadsheets")),
    )

    text = await DoclingAdapter()._process_single_file(str(path), path.name)

    assert text == "compact values\n"


@pytest.mark.asyncio
async def test_pdf_with_dense_private_use_text_retries_with_full_page_ocr(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "encoded.pdf"
    path.write_bytes(b"%PDF encoded font")
    encoded = "".join(chr(0xE100 + index) for index in range(20))

    @asynccontextmanager
    async def slot(*_args, **_kwargs):
        yield None

    monkeypatch.setattr("lib.services_gateway.gateway.slot", slot)
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_sync",
        staticmethod(lambda _path: encoded),
    )
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_force_ocr_sync",
        staticmethod(lambda _path: "Readable OCR text."),
    )

    text = await DoclingAdapter()._process_single_file(str(path), path.name)

    assert text == "Readable OCR text."


@pytest.mark.asyncio
async def test_pdf_fails_when_full_page_ocr_remains_private_use(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "encoded.pdf"
    path.write_bytes(b"%PDF encoded font")
    encoded = "".join(chr(0xE100 + index) for index in range(20))

    @asynccontextmanager
    async def slot(*_args, **_kwargs):
        yield None

    monkeypatch.setattr("lib.services_gateway.gateway.slot", slot)
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_sync",
        staticmethod(lambda _path: encoded),
    )
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_force_ocr_sync",
        staticmethod(lambda _path: encoded),
    )

    with pytest.raises(RuntimeError, match="Full-page OCR still produced"):
        await DoclingAdapter()._process_single_file(
            str(path),
            path.name,
            raise_on_error=True,
        )


def test_repaired_pdf_conversion_runs_ghostscript_then_docling(monkeypatch, tmp_path):
    source = tmp_path / "source.pdf"
    source.write_bytes(b"%PDF-1.4\n%%EOF\n")
    calls = {}

    monkeypatch.setattr(
        "lib.adapters.docling.pdf.shutil.which",
        lambda name: "/usr/bin/gs",
    )

    def fake_run(cmd, check, stdout, stderr):
        calls["cmd"] = cmd
        output_arg = next(item for item in cmd if item.startswith("-sOutputFile="))
        repaired = output_arg.split("=", 1)[1]
        assert cmd[-1] == str(source)
        with open(repaired, "wb") as f:
            f.write(b"%PDF-1.4 repaired\n%%EOF\n")

    monkeypatch.setattr("lib.adapters.docling.pdf.subprocess.run", fake_run)
    monkeypatch.setattr(
        DoclingAdapter,
        "_convert_sync",
        staticmethod(lambda path: f"converted {path.endswith('.pdf')}"),
    )

    markdown = DoclingAdapter._convert_repaired_pdf_sync(str(source))

    assert markdown == "converted True"
    assert calls["cmd"][0] == "/usr/bin/gs"
    assert "-sDEVICE=pdfwrite" in calls["cmd"]


def test_chat_completions_url_does_not_duplicate_v1():
    assert _chat_completions_url("http://localhost:8080/v1") == (
        "http://localhost:8080/v1/chat/completions"
    )
    assert _chat_completions_url("http://localhost:11434") == (
        "http://localhost:11434/v1/chat/completions"
    )


def test_chat_completions_model_strips_ollama_prefix_only_for_ollama_host():
    assert _chat_completions_model("http://localhost:11434", "ollama/qwen3-vl:8b") == (
        "qwen3-vl:8b"
    )
    assert _chat_completions_model("http://localhost:8080/v1", "ollama/qwen3-vl:8b") == (
        "ollama/qwen3-vl:8b"
    )


def test_picture_description_params_use_max_completion_tokens():
    from lib.adapters.docling.converter import picture_description_params

    assert picture_description_params("gpt-5.6-luna") == {
        "model": "gpt-5.6-luna",
        "max_completion_tokens": 200,
    }
    assert "max_tokens" not in picture_description_params("gpt-4o-mini")
