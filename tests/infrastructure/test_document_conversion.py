from __future__ import annotations

import inspect

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from lib.infrastructure.document_conversion import (
    DocumentConversion,
    SPREADSHEET_CONVERSION_MARKER,
    convert_document,
)
from lib.infrastructure.document_conversion.docling_stack import (
    converter as stack_converter,
)
from lib.infrastructure.document_conversion.docling_stack.docling import (
    chat_completions_model,
    chat_completions_url,
    picture_description_params,
)
from lib.infrastructure.document_conversion.docling_stack.spreadsheets import (
    convert_spreadsheet,
)
from lib.infrastructure.errors import (
    InfrastructureError,
    InfrastructureErrorKind,
)


def test_docling_sdk_import_available():
    from docling.document_converter import DocumentConverter

    assert DocumentConverter is not None


@pytest.mark.asyncio
async def test_rtf_conversion_extracts_text_without_docling(tmp_path):
    path = tmp_path / "readme.rtf"
    path.write_text(
        r"{\rtf1\ansi\ansicpg1252 This folder contains:\par Confidential caf\'e9.}",
        encoding="latin-1",
    )

    conversion = await stack_converter.convert_document(path)

    assert conversion.markdown == "This folder contains:\nConfidential café.\n"


@pytest.mark.asyncio
async def test_rtf_processing_bypasses_docling(monkeypatch, tmp_path):
    path = tmp_path / "readme.rtf"
    path.write_text(r"{\rtf1\ansi Plain text.}", encoding="latin-1")
    monkeypatch.setattr(
        stack_converter,
        "convert_with_docling",
        lambda _path: pytest.fail("Docling should not handle RTF"),
    )

    conversion = await stack_converter.convert_document(path)

    assert conversion.markdown == "Plain text.\n"


@pytest.mark.asyncio
async def test_public_interface_returns_empty_source_warning(tmp_path):
    path = tmp_path / "empty.pdf"
    path.write_bytes(b"")

    conversion = await convert_document(path)

    assert conversion.markdown == ""
    assert conversion.warnings == ("The source file is empty",)


@pytest.mark.asyncio
async def test_public_interface_schedules_actual_docling_call(
    mocker,
    tmp_path,
):
    path = tmp_path / "report.pdf"
    path.write_bytes(b"%PDF fixture")
    docling = mocker.patch.object(
        stack_converter,
        "convert_with_docling",
        return_value="Converted report",
    )
    calls = []

    async def run_now(operation, **kwargs):
        calls.append(kwargs)
        result = operation(**kwargs["operation_kwargs"])
        return await result if inspect.isawaitable(result) else result

    scheduler_run = mocker.patch.object(
        stack_converter.scheduler,
        "run",
        side_effect=run_now,
    )

    conversion = await convert_document(path)

    assert conversion.markdown == "Converted report"
    docling.assert_called_once_with(str(path))
    scheduler_run.assert_awaited_once()
    profile = stack_converter._inspect_docling(
        calls[0]["operation_kwargs"]
    )
    assert profile.descriptor == "docling"
    assert profile.input_size == path.stat().st_size


@pytest.mark.asyncio
async def test_public_interface_does_not_schedule_non_docling_conversion(
    mocker,
    tmp_path,
):
    path = tmp_path / "notes.txt"
    path.write_text("Plain text", encoding="utf-8")
    scheduler_run = mocker.patch.object(stack_converter.scheduler, "run")

    conversion = await convert_document(path)

    assert conversion.markdown == "Plain text"
    scheduler_run.assert_not_awaited()


@pytest.mark.asyncio
async def test_public_interface_wraps_unsupported_formats(tmp_path):
    path = tmp_path / "logo.eps"
    path.write_bytes(b"%!PS-Adobe EPSF")

    with pytest.raises(InfrastructureError) as raised:
        await convert_document(path)

    assert raised.value.kind is InfrastructureErrorKind.INVALID_RESPONSE
    assert raised.value.operation == "check_format"


@pytest.mark.asyncio
async def test_public_interface_rejects_unknown_provider(monkeypatch, tmp_path):
    path = tmp_path / "readme.txt"
    path.write_text("text", encoding="utf-8")
    monkeypatch.setenv("DOCUMENT_CONVERTER", "unknown")

    with pytest.raises(InfrastructureError) as raised:
        await convert_document(path)

    assert raised.value.kind is InfrastructureErrorKind.CONFIGURATION
    assert raised.value.operation == "select_provider"


@pytest.mark.asyncio
async def test_public_interface_wraps_provider_errors(monkeypatch, tmp_path):
    from lib.infrastructure.document_conversion import converter

    path = tmp_path / "broken.pdf"
    path.write_bytes(b"%PDF broken")

    def fail(_path):
        raise RuntimeError("converter unavailable")

    monkeypatch.setattr(converter, "_backend", lambda _provider: fail)

    with pytest.raises(InfrastructureError) as raised:
        await convert_document(path)

    assert raised.value.kind is InfrastructureErrorKind.INVALID_RESPONSE
    assert raised.value.operation == "convert_document"
    assert "converter unavailable" in str(raised.value)


def test_spreadsheet_conversion_is_compact_valid_markdown(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "financial model"
    sheet["A1"] = "Metric"
    sheet["B1"] = "Value"
    sheet["A2"] = "Revenue"
    sheet["B2"] = 1200
    sheet.merge_cells("A5:XFD5")
    path = tmp_path / "model.xlsx"
    workbook.save(path)

    conversion = convert_spreadsheet(path)

    assert conversion.markdown.startswith(SPREADSHEET_CONVERSION_MARKER)
    assert "## financial model" in conversion.markdown
    assert "| Metric | Value |" in conversion.markdown
    assert "| --- | --- |" in conversion.markdown
    assert "| Revenue | 1200 |" in conversion.markdown
    assert len(conversion.markdown) < 1_000
    assert "XFD" not in conversion.markdown


def test_spreadsheet_conversion_warns_about_uncached_formulas(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Amount"
    sheet["A2"] = "Seed"
    sheet["B2"] = "=40+60"
    path = tmp_path / "normal.xlsx"
    workbook.save(path)

    conversion = convert_spreadsheet(path)

    assert "Seed" in conversion.markdown
    assert "=40+60" not in conversion.markdown
    assert conversion.warnings == (
        "Omitted 1 formula cell(s) without cached values",
    )


def test_spreadsheet_conversion_pads_rows_to_a_valid_table_width(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Amount"])
    sheet.append(["Seed"])
    path = tmp_path / "uneven.xlsx"
    workbook.save(path)

    conversion = convert_spreadsheet(path)

    assert "| Name | Amount |" in conversion.markdown
    assert "| Seed |  |" in conversion.markdown


def test_spreadsheet_conversion_omits_hidden_content_and_errors(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "visible"
    sheet["A1"] = "Revenue"
    sheet["A2"] = "hidden row"
    sheet.row_dimensions[2].hidden = True
    sheet["B1"] = "hidden column"
    sheet.column_dimensions["B"].hidden = True
    sheet["C1"] = "#DIV/0!"
    sheet["C1"].data_type = "e"
    hidden_sheet = workbook.create_sheet("hidden sheet")
    hidden_sheet.sheet_state = "hidden"
    hidden_sheet["A1"] = "hidden worksheet"
    path = tmp_path / "hidden-content.xlsx"
    workbook.save(path)

    conversion = convert_spreadsheet(path)

    assert "Revenue" in conversion.markdown
    assert "hidden row" not in conversion.markdown
    assert "hidden column" not in conversion.markdown
    assert "#DIV/0!" not in conversion.markdown
    assert "hidden worksheet" not in conversion.markdown
    assert conversion.warnings == ("Omitted 1 Excel error cell(s)",)


def test_spreadsheet_conversion_reports_unconverted_visuals(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Month", "Revenue"])
    sheet.append(["Jan", 100])
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=2))
    sheet.add_chart(chart, "D2")
    path = tmp_path / "chart.xlsx"
    workbook.save(path)

    conversion = convert_spreadsheet(path)

    assert conversion.warnings == (
        "Did not convert 1 embedded chart(s) and 0 embedded image(s)",
    )


@pytest.mark.parametrize("extension", [".xls", ".xlsx", ".xlsm"])
@pytest.mark.asyncio
async def test_spreadsheet_processing_bypasses_docling(
    monkeypatch,
    tmp_path,
    extension,
):
    path = tmp_path / f"model{extension}"
    path.write_bytes(b"workbook")
    monkeypatch.setattr(
        stack_converter,
        "convert_spreadsheet",
        lambda _path: DocumentConversion(markdown="compact values\n"),
    )
    monkeypatch.setattr(
        stack_converter,
        "convert_with_docling",
        lambda _path: pytest.fail("Docling should not handle spreadsheets"),
    )

    conversion = await stack_converter.convert_document(path)

    assert conversion.markdown == "compact values\n"


@pytest.mark.asyncio
async def test_pdf_with_dense_private_use_text_retries_with_full_page_ocr(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "encoded.pdf"
    path.write_bytes(b"%PDF encoded font")
    encoded = "".join(chr(0xE100 + index) for index in range(20))
    monkeypatch.setattr(
        stack_converter,
        "convert_with_docling",
        lambda _path: encoded,
    )
    monkeypatch.setattr(
        stack_converter,
        "convert_document_force_ocr",
        lambda _path: "Readable OCR text.",
    )
    scheduled_calls = []

    async def run_now(operation, **_kwargs):
        scheduled_calls.append(operation)
        result = operation(**_kwargs["operation_kwargs"])
        return await result if inspect.isawaitable(result) else result

    monkeypatch.setattr(stack_converter.scheduler, "run", run_now)

    conversion = await stack_converter.convert_document(path)

    assert conversion.markdown == "Readable OCR text."
    assert conversion.warnings == ("Full-page OCR was required",)
    assert len(scheduled_calls) == 2


@pytest.mark.asyncio
async def test_pdf_fails_when_full_page_ocr_remains_private_use(
    monkeypatch,
    tmp_path,
):
    path = tmp_path / "encoded.pdf"
    path.write_bytes(b"%PDF encoded font")
    encoded = "".join(chr(0xE100 + index) for index in range(20))
    monkeypatch.setattr(
        stack_converter,
        "convert_with_docling",
        lambda _path: encoded,
    )
    monkeypatch.setattr(
        stack_converter,
        "convert_document_force_ocr",
        lambda _path: encoded,
    )

    async def run_now(operation, **_kwargs):
        result = operation(**_kwargs["operation_kwargs"])
        return await result if inspect.isawaitable(result) else result

    monkeypatch.setattr(stack_converter.scheduler, "run", run_now)

    with pytest.raises(RuntimeError, match="Full-page OCR still produced"):
        await stack_converter.convert_document(path)


def test_repaired_pdf_conversion_runs_ghostscript_then_docling(
    monkeypatch,
    tmp_path,
):
    from lib.infrastructure.document_conversion.docling_stack import pdf

    source = tmp_path / "source.pdf"
    source.write_bytes(b"%PDF-1.4\n%%EOF\n")
    calls = {}
    monkeypatch.setattr(pdf.shutil, "which", lambda name: "/usr/bin/gs")

    def fake_run(cmd, check, stdout, stderr):
        calls["cmd"] = cmd
        output_arg = next(item for item in cmd if item.startswith("-sOutputFile="))
        repaired = output_arg.split("=", 1)[1]
        assert cmd[-1] == str(source)
        with open(repaired, "wb") as handle:
            handle.write(b"%PDF-1.4 repaired\n%%EOF\n")

    monkeypatch.setattr(pdf.subprocess, "run", fake_run)

    markdown = pdf.convert_repaired_pdf(
        str(source),
        lambda path: f"converted {path.endswith('.pdf')}",
    )

    assert markdown == "converted True"
    assert calls["cmd"][0] == "/usr/bin/gs"
    assert "-sDEVICE=pdfwrite" in calls["cmd"]


def test_chat_completions_url_does_not_duplicate_v1():
    assert chat_completions_url("http://localhost:8080/v1") == (
        "http://localhost:8080/v1/chat/completions"
    )
    assert chat_completions_url("http://localhost:11434") == (
        "http://localhost:11434/v1/chat/completions"
    )


def test_chat_completions_model_strips_ollama_prefix_only_for_ollama_host():
    assert chat_completions_model(
        "http://localhost:11434",
        "ollama/qwen3-vl:8b",
    ) == "qwen3-vl:8b"
    assert chat_completions_model(
        "http://localhost:8080/v1",
        "ollama/qwen3-vl:8b",
    ) == "ollama/qwen3-vl:8b"


def test_picture_description_params_use_max_completion_tokens():
    assert picture_description_params("gpt-5.6-luna") == {
        "model": "gpt-5.6-luna",
        "max_completion_tokens": 200,
    }
    assert "max_tokens" not in picture_description_params("gpt-4o-mini")
