from __future__ import annotations

import datetime

from lib.datasets.spreadsheet_markdown import SPREADSHEET_MARKDOWN_MARKER
from lib.logger import get_logger

logger = get_logger(__name__)

SPREADSHEET_EXTENSIONS = (".xls", ".xlsx", ".xlsm")

__all__ = [
    "SPREADSHEET_EXTENSIONS",
    "SPREADSHEET_MARKDOWN_MARKER",
    "convert_spreadsheet",
    "format_cell_value",
    "is_spreadsheet_filename",
]


def is_spreadsheet_filename(filename: str) -> bool:
    return filename.lower().endswith(SPREADSHEET_EXTENSIONS)


def convert_spreadsheet(filepath: str) -> str:
    """Convert spreadsheets to compact, value-only Markdown."""
    if filepath.lower().endswith(".xls"):
        return convert_xls(filepath)
    return convert_openpyxl(filepath)


def convert_openpyxl(filepath: str) -> str:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    values_wb = load_workbook(filepath, read_only=False, data_only=True)
    formulas_wb = load_workbook(filepath, read_only=False, data_only=False)
    sections = []
    missing_cached_formulas = 0
    try:
        for values_ws, formulas_ws in zip(
            values_wb.worksheets,
            formulas_wb.worksheets,
        ):
            if values_ws.sheet_state != "visible":
                continue

            def is_hidden(row: int, column: int) -> bool:
                row_dimension = values_ws.row_dimensions.get(row)
                if row_dimension is not None and row_dimension.hidden:
                    return True
                column_dimension = values_ws.column_dimensions.get(
                    get_column_letter(column)
                )
                return bool(
                    column_dimension is not None
                    and column_dimension.hidden
                )

            row_cells: dict[int, dict[int, str]] = {}
            for (row, col), cell in values_ws._cells.items():
                if is_hidden(row, col) or cell.data_type == "e":
                    continue
                text = format_cell_value(
                    cell.value,
                    getattr(cell, "number_format", ""),
                )
                if text:
                    row_cells.setdefault(row, {})[col] = text
            for (row, col), formula_cell in formulas_ws._cells.items():
                if is_hidden(row, col):
                    continue
                cached_cell = values_ws._cells.get((row, col))
                if (
                    formula_cell.data_type == "f"
                    and (
                        cached_cell is None
                        or cached_cell.value in (None, "")
                    )
                ):
                    missing_cached_formulas += 1

            if not row_cells:
                continue
            sections.append(f"## {values_ws.title}")
            for row_index in sorted(row_cells):
                columns = row_cells[row_index]
                row = [
                    columns.get(column_index, "")
                    for column_index in range(1, max(columns) + 1)
                ]
                sections.append(
                    "| "
                    + " | ".join(
                        escape_markdown_cell(cell) for cell in row
                    )
                    + " |"
                )
            sections.append("")
    finally:
        values_wb.close()
        formulas_wb.close()

    if missing_cached_formulas:
        logger.warning(
            "Spreadsheet conversion omitted %s formula cells without "
            "cached values: %s",
            missing_cached_formulas,
            filepath,
        )
    return render_spreadsheet_markdown(sections)


def convert_xls(filepath: str) -> str:
    import xlrd

    workbook = xlrd.open_workbook(filepath, on_demand=True)
    sections = []
    try:
        for sheet in workbook.sheets():
            rows = []
            for row_index in range(sheet.nrows):
                row_info = sheet.rowinfo_map.get(row_index)
                if row_info is not None and row_info.hidden:
                    continue
                values = [
                    ""
                    if (
                        (column_info := sheet.colinfo_map.get(column_index))
                        is not None
                        and column_info.hidden
                    )
                    else xls_cell_text(
                        sheet.cell_value(row_index, column_index)
                    )
                    for column_index in range(sheet.ncols)
                ]
                while values and not values[-1]:
                    values.pop()
                if any(values):
                    rows.append(values)
            if not rows:
                continue
            sections.append(f"## {sheet.name}")
            sections.extend(
                "| "
                + " | ".join(
                    escape_markdown_cell(value) for value in row
                )
                + " |"
                for row in rows
            )
            sections.append("")
    finally:
        workbook.release_resources()
    return render_spreadsheet_markdown(sections)


def escape_markdown_cell(value: str) -> str:
    return value.replace("|", "\\|")


def format_cell_value(value, number_format: str = "") -> str:
    """Render a cell as text a reader (and an embedding model) can use."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return format_temporal(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format_float(value, number_format)
    return str(value).replace("\n", " ").strip()


def format_temporal(value) -> str:
    if isinstance(value, datetime.datetime):
        if (value.hour, value.minute, value.second) == (0, 0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, datetime.time):
        return value.isoformat(timespec="minutes")
    return value.isoformat()


def format_float(value: float, number_format: str = "") -> str:
    if value != value or value in (float("inf"), float("-inf")):
        return str(value)
    if "%" in (number_format or ""):
        return f"{format_magnitude(value * 100)}%"
    return format_magnitude(value)


def format_magnitude(value: float) -> str:
    """Drop float noise: spreadsheet ratios carry far more digits than meaning."""
    if value == int(value):
        return str(int(value))
    magnitude = abs(value)
    if magnitude >= 1:
        text = f"{value:.2f}"
    elif magnitude >= 1e-4:
        text = f"{value:.4f}"
    else:
        return f"{value:.4g}"
    return text.rstrip("0").rstrip(".")


def render_spreadsheet_markdown(sections: list[str]) -> str:
    body = "\n".join(sections).strip()
    if not body:
        return f"{SPREADSHEET_MARKDOWN_MARKER}\n"
    return f"{SPREADSHEET_MARKDOWN_MARKER}\n\n{body}\n"


def xls_cell_text(value) -> str:
    if value is None or value == "":
        return ""
    return format_cell_value(value)
