"""
DoclingAdapter — thin wrapper around the docling library for OCR/parsing.

Runs in-process. No HTTP, no separate server. Models load on first call and
are reused for the lifetime of the process via a module-level singleton.

OCR backend: Apple Vision via ocrmac on macOS; RapidOCR on Linux.
Picture descriptions are forwarded to a configured VLM service.
"""
import asyncio
import os
import platform
import shutil
import subprocess
import tempfile
import threading
from typing import List
from urllib.parse import urlparse

from lib.env import get_env_var
from lib.logger import get_logger

logger = get_logger(__name__)

_PASSTHROUGH_EXTS = (".json", ".txt", ".md")
_RTF_EXTS = (".rtf",)
_SPREADSHEET_EXTS = (".xls", ".xlsx", ".xlsm")
SPREADSHEET_MARKDOWN_MARKER = "<!-- sictic-spreadsheet: compact-values-v1 -->"

_converter = None
_converter_init_lock = threading.Lock()
# docling's pipeline objects aren't documented as thread-safe; serialise convert() calls.
_convert_lock = threading.Lock()


def _build_converter():
    """Construct the DocumentConverter — called once, lazily."""
    from lib.runtime_noise import configure_runtime_noise

    configure_runtime_noise()

    from docling.document_converter import DocumentConverter, PdfFormatOption
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import (
        PdfPipelineOptions,
        PictureDescriptionApiOptions,
        OcrMacOptions,
        RapidOcrOptions,
    )

    vlm_model_name = get_env_var("VLM_MODEL")
    vlm_base_url = (
        os.environ.get("VLM_BASE_URL")
        or os.environ.get("LLM_BASE_URL")
        or os.environ.get("OLLAMA_HOST")
        or "http://localhost:11434"
    ).rstrip("/")
    vlm_model_name = _chat_completions_model(vlm_base_url, vlm_model_name)
    vlm_api_key = os.environ.get("VLM_API_KEY") or os.environ.get("LLM_API_KEY") or ""
    headers = {"Authorization": f"Bearer {vlm_api_key}"} if vlm_api_key else {}

    if platform.system() == "Darwin":
        ocr_options = OcrMacOptions()
    else:
        ocr_options = RapidOcrOptions()

    pipeline_opts = PdfPipelineOptions(
        do_ocr=True,
        do_picture_description=True,
        enable_remote_services=True,
        ocr_options=ocr_options,
        picture_description_options=PictureDescriptionApiOptions(
            url=_chat_completions_url(vlm_base_url),
            headers=headers,
            params={"model": vlm_model_name, "max_tokens": 200},
            prompt="Describe this image in a few sentences.",
            timeout=600,
        ),
    )
    return DocumentConverter(
        format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_opts)}
    )


def _get_converter():
    global _converter
    if _converter is not None:
        return _converter
    with _converter_init_lock:
        if _converter is None:
            logger.info("Initializing docling DocumentConverter (first use; models load on first convert)")
            _converter = _build_converter()
    return _converter


class DoclingAdapter:
    def __init__(self, concurrency_limit: int = 10):
        self.concurrency_limit = concurrency_limit

    async def extract_documents(self, files_to_process: List[dict]):
        """Yield (filename, markdown) for each file as conversion finishes.

        Each item in files_to_process must include:
          - "filename": the logical name used in returned tuples and downstream paths
          - "local_path": an on-disk absolute path docling can open (use
            storage.local_path(rel) to materialize a Drive file locally first).
        """
        tasks = []
        for f_data in files_to_process:
            filename = f_data["filename"]
            filepath = str(f_data["local_path"])

            async def run(fname=filename, fp=filepath):
                txt = await self._process_single_file(fp, fname)
                return fname, txt

            tasks.append(asyncio.create_task(run()))

        for coro in asyncio.as_completed(tasks):
            yield await coro

    async def _process_single_file(self, filepath: str, filename: str) -> str:
        from lib.services_gateway import gateway

        await gateway.acquire_docling_slot(self.concurrency_limit)
        try:
            if not os.path.isfile(filepath):
                logger.warning(f"Skipping {filepath}, not a valid local file.")
                return ""
            if os.path.getsize(filepath) == 0:
                logger.warning(f"Skipping {filepath}, file is empty (0 bytes).")
                return ""

            if filename.lower().endswith(_PASSTHROUGH_EXTS):
                with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                    return f.read()
            if filename.lower().endswith(_RTF_EXTS):
                return await asyncio.to_thread(self._convert_rtf_sync, filepath)
            if is_spreadsheet_filename(filename):
                return await asyncio.to_thread(self._convert_spreadsheet_sync, filepath)

            try:
                return await asyncio.to_thread(self._convert_sync, filepath)
            except Exception as e:
                if filename.lower().endswith(".pdf") and "page-dimensions" in str(e):
                    logger.warning(
                        f"Docling conversion failed for {filename}: could not resolve page dimensions. "
                        "Retrying after Ghostscript PDF normalization."
                    )
                    try:
                        return await asyncio.to_thread(self._convert_repaired_pdf_sync, filepath)
                    except Exception as repair_error:
                        logger.error(
                            f"Docling conversion failed for {filename} after Ghostscript normalization: "
                            f"{_short_error(repair_error)}"
                        )
                        return ""
                logger.error(f"Docling conversion failed for {filename}: {_short_error(e)}")
                return ""
        finally:
            gateway.release_docling_slot()

    @staticmethod
    def _convert_sync(filepath: str) -> str:
        converter = _get_converter()
        with _convert_lock:
            result = converter.convert(filepath)
        return result.document.export_to_markdown()

    @staticmethod
    def _convert_rtf_sync(filepath: str) -> str:
        """Extract searchable text from RTF, which Docling does not support."""
        from striprtf.striprtf import rtf_to_text

        with open(filepath, "r", encoding="latin-1") as handle:
            rtf = handle.read()
        text = rtf_to_text(rtf, errors="replace")
        return text.strip() + "\n" if text.strip() else ""

    @staticmethod
    def _convert_repaired_pdf_sync(filepath: str) -> str:
        gs = shutil.which("gs")
        if not gs:
            raise RuntimeError("Ghostscript executable 'gs' not found.")
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            repaired_path = tmp.name
        try:
            subprocess.run(
                [
                    gs,
                    "-q",
                    "-dNOPAUSE",
                    "-dBATCH",
                    "-sDEVICE=pdfwrite",
                    "-dCompatibilityLevel=1.4",
                    f"-sOutputFile={repaired_path}",
                    filepath,
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            return DoclingAdapter._convert_sync(repaired_path)
        finally:
            try:
                os.remove(repaired_path)
            except FileNotFoundError:
                pass

    @staticmethod
    def _convert_spreadsheet_sync(filepath: str) -> str:
        """Convert spreadsheets to compact, value-only Markdown."""
        if filepath.lower().endswith(".xls"):
            return DoclingAdapter._convert_xls_sync(filepath)
        return DoclingAdapter._convert_openpyxl_sync(filepath)

    @staticmethod
    def _convert_openpyxl_sync(filepath: str) -> str:
        from openpyxl import load_workbook

        values_wb = load_workbook(filepath, read_only=False, data_only=True)
        formulas_wb = load_workbook(filepath, read_only=False, data_only=False)
        sections = []
        missing_cached_formulas = 0

        for values_ws, formulas_ws in zip(values_wb.worksheets, formulas_wb.worksheets):
            row_cells: dict[int, dict[int, str]] = {}
            for row, col in values_ws._cells:
                value = values_ws.cell(row=row, column=col).value
                text = "" if value is None else str(value).replace("\n", " ").strip()
                if text:
                    row_cells.setdefault(row, {})[col] = text
            for row, col in formulas_ws._cells:
                formula_cell = formulas_ws.cell(row=row, column=col)
                if (
                    formula_cell.data_type == "f"
                    and values_ws.cell(row=row, column=col).value in (None, "")
                ):
                    missing_cached_formulas += 1

            if not row_cells:
                continue

            sections.append(f"## {values_ws.title}")
            for row_idx in sorted(row_cells):
                cols = row_cells[row_idx]
                max_col = max(cols)
                row = [cols.get(col_idx, "") for col_idx in range(1, max_col + 1)]
                sections.append("| " + " | ".join(_escape_markdown_cell(c) for c in row) + " |")
            sections.append("")

        values_wb.close()
        formulas_wb.close()
        if missing_cached_formulas:
            logger.warning(
                "Spreadsheet conversion omitted %s formula cells without cached values: %s",
                missing_cached_formulas,
                filepath,
            )
        return _render_spreadsheet_markdown(sections)

    @staticmethod
    def _convert_xls_sync(filepath: str) -> str:
        import xlrd

        workbook = xlrd.open_workbook(filepath, on_demand=True)
        sections = []
        try:
            for sheet in workbook.sheets():
                rows = []
                for row_idx in range(sheet.nrows):
                    values = [
                        _xls_cell_text(sheet.cell_value(row_idx, col_idx))
                        for col_idx in range(sheet.ncols)
                    ]
                    while values and not values[-1]:
                        values.pop()
                    if any(values):
                        rows.append(values)
                if not rows:
                    continue
                sections.append(f"## {sheet.name}")
                sections.extend(
                    "| " + " | ".join(_escape_markdown_cell(value) for value in row) + " |"
                    for row in rows
                )
                sections.append("")
        finally:
            workbook.release_resources()
        return _render_spreadsheet_markdown(sections)


def _escape_markdown_cell(value: str) -> str:
    return value.replace("|", "\\|")


def is_spreadsheet_filename(filename: str) -> bool:
    return filename.lower().endswith(_SPREADSHEET_EXTS)


def _render_spreadsheet_markdown(sections: list[str]) -> str:
    body = "\n".join(sections).strip()
    if not body:
        return f"{SPREADSHEET_MARKDOWN_MARKER}\n"
    return f"{SPREADSHEET_MARKDOWN_MARKER}\n\n{body}\n"


def _xls_cell_text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\n", " ").strip()


def _short_error(error: Exception, limit: int = 500) -> str:
    text = str(error).replace("\n", " ").strip()
    if len(text) > limit:
        return text[:limit].rstrip() + "..."
    return text


def _chat_completions_url(base_url: str) -> str:
    base = base_url.rstrip("/")
    if base.endswith("/v1"):
        return f"{base}/chat/completions"
    return f"{base}/v1/chat/completions"


def _chat_completions_model(base_url: str, model: str) -> str:
    if model.startswith("ollama/") and _is_ollama_base_url(base_url):
        return model[7:]
    return model


def _is_ollama_base_url(base_url: str) -> bool:
    parsed = urlparse(base_url)
    host = parsed.hostname or ""
    return host in {"localhost", "127.0.0.1", "::1"} and parsed.port == 11434
