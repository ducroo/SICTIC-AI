from __future__ import annotations

import datetime
from pathlib import Path

from lib.infrastructure.document_conversion.types import (
    DocumentConversion,
    SPREADSHEET_CONVERSION_MARKER,
)
from lib.markdown_tables import select_header

SPREADSHEET_EXTENSIONS = (".xls", ".xlsx", ".xlsm")

__all__ = [
    "SPREADSHEET_EXTENSIONS",
    "convert_spreadsheet",
    "format_cell_value",
    "is_spreadsheet_filename",
]


def is_spreadsheet_filename(filename: str) -> bool:
    return filename.lower().endswith(SPREADSHEET_EXTENSIONS)


def convert_spreadsheet(filepath: str | Path) -> DocumentConversion:
    """Convert spreadsheets to compact, value-only Markdown."""
    path = Path(filepath)
    if path.suffix.lower() == ".xls":
        return DocumentConversion(markdown=convert_xls(str(path)))
    return convert_openpyxl(str(path))


def convert_openpyxl(filepath: str) -> DocumentConversion:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    values_wb = load_workbook(filepath, read_only=False, data_only=True)
    formulas_wb = load_workbook(filepath, read_only=False, data_only=False)
    sections = []
    missing_cached_formulas = 0
    ignored_error_cells = 0
    ignored_charts = 0
    ignored_images = 0
    try:
        for values_ws, formulas_ws in zip(
            values_wb.worksheets,
            formulas_wb.worksheets,
        ):
            if values_ws.sheet_state != "visible":
                continue
            ignored_charts += len(values_ws._charts)
            ignored_images += len(values_ws._images)

            def is_hidden(row: int, column: int) -> bool:
                row_dimension = values_ws.row_dimensions.get(row)
                if row_dimension is not None and row_dimension.hidden:
                    return True
                column_dimension = values_ws.column_dimensions.get(
                    get_column_letter(column)
                )
                return bool(
                    column_dimension is not None
                    and column_dimension.hidden
                )

            row_cells: dict[int, dict[int, str]] = {}
            for (row, col), cell in values_ws._cells.items():
                if is_hidden(row, col):
                    continue
                if cell.data_type == "e":
                    ignored_error_cells += 1
                    continue
                text = format_cell_value(
                    cell.value,
                    getattr(cell, "number_format", ""),
                )
                if text:
                    row_cells.setdefault(row, {})[col] = text
            for (row, col), formula_cell in formulas_ws._cells.items():
                if is_hidden(row, col):
                    continue
                cached_cell = values_ws._cells.get((row, col))
                if (
                    formula_cell.data_type == "f"
                    and (
                        cached_cell is None
                        or cached_cell.value in (None, "")
                    )
                ):
                    missing_cached_formulas += 1

            if not row_cells:
                continue
            sections.append(f"## {values_ws.title}")
            rows: list[list[str]] = []
            for row_index in sorted(row_cells):
                columns = row_cells[row_index]
                rows.append(
                    [
                        columns.get(column_index, "")
                        for column_index in range(1, max(columns) + 1)
                    ]
                )
            sections.extend(render_rows(rows))
            sections.append("")
    finally:
        values_wb.close()
        formulas_wb.close()

    warnings: list[str] = []
    if missing_cached_formulas:
        warnings.append(
            f"Omitted {missing_cached_formulas} formula cell(s) without "
            "cached values"
        )
    if ignored_error_cells:
        warnings.append(
            f"Omitted {ignored_error_cells} Excel error cell(s)"
        )
    if ignored_charts or ignored_images:
        warnings.append(
            f"Did not convert {ignored_charts} embedded chart(s) and "
            f"{ignored_images} embedded image(s)"
        )
    return DocumentConversion(
        markdown=render_spreadsheet_markdown(sections),
        warnings=tuple(warnings),
    )


def convert_xls(filepath: str) -> str:
    import xlrd

    workbook = xlrd.open_workbook(filepath, on_demand=True)
    sections = []
    try:
        for sheet in workbook.sheets():
            rows = []
            for row_index in range(sheet.nrows):
                row_info = sheet.rowinfo_map.get(row_index)
                if row_info is not None and row_info.hidden:
                    continue
                values = [
                    ""
                    if (
                        (column_info := sheet.colinfo_map.get(column_index))
                        is not None
                        and column_info.hidden
                    )
                    else xls_cell_text(
                        sheet.cell_value(row_index, column_index)
                    )
                    for column_index in range(sheet.ncols)
                ]
                while values and not values[-1]:
                    values.pop()
                if any(values):
                    rows.append(values)
            if not rows:
                continue
            sections.append(f"## {sheet.name}")
            sections.extend(render_rows(rows))
            sections.append("")
    finally:
        workbook.release_resources()
    return render_spreadsheet_markdown(sections)


def escape_markdown_cell(value: str) -> str:
    return value.replace("|", "\\|")


def render_rows(rows: list[list[str]]) -> list[str]:
    """Render one worksheet as compact, valid Markdown."""
    if not rows:
        return []
    rendered = [render_row(row) for row in rows]
    header_index = select_header(rendered)
    if header_index < 0:
        width = max(len(row) for row in rows)
        return [
            render_row([""] * width),
            render_separator(width),
            *(render_row(pad_row(row, width)) for row in rows),
        ]

    prefix = [render_prose_row(row) for row in rows[:header_index]]
    table_rows = rows[header_index:]
    width = max(len(row) for row in table_rows)
    return [
        *prefix,
        render_row(pad_row(table_rows[0], width)),
        render_separator(width),
        *(render_row(pad_row(row, width)) for row in table_rows[1:]),
    ]


def pad_row(row: list[str], width: int) -> list[str]:
    return row + [""] * (width - len(row))


def render_prose_row(row: list[str]) -> str:
    values = [escape_markdown_cell(value) for value in row if value]
    return " — ".join(values)


def render_row(row: list[str]) -> str:
    return (
        "| "
        + " | ".join(escape_markdown_cell(cell) for cell in row)
        + " |"
    )


def render_separator(columns: int) -> str:
    return "| " + " | ".join("---" for _ in range(columns)) + " |"


def format_cell_value(value, number_format: str = "") -> str:
    """Render a cell as text a reader (and an embedding model) can use."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return format_temporal(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format_float(value, number_format)
    return str(value).replace("\n", " ").strip()


def format_temporal(value) -> str:
    if isinstance(value, datetime.datetime):
        if (value.hour, value.minute, value.second) == (0, 0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, datetime.time):
        return value.isoformat(timespec="minutes")
    return value.isoformat()


def format_float(value: float, number_format: str = "") -> str:
    if value != value or value in (float("inf"), float("-inf")):
        return str(value)
    if "%" in (number_format or ""):
        return f"{format_magnitude(value * 100)}%"
    return format_magnitude(value)


def format_magnitude(value: float) -> str:
    """Drop float noise: spreadsheet ratios carry far more digits than meaning."""
    if value == int(value):
        return str(int(value))
    magnitude = abs(value)
    if magnitude >= 1:
        text = f"{value:.2f}"
    elif magnitude >= 1e-4:
        text = f"{value:.4f}"
    else:
        return f"{value:.4g}"
    return text.rstrip("0").rstrip(".")


def render_spreadsheet_markdown(sections: list[str]) -> str:
    body = "\n".join(sections).strip()
    if not body:
        return f"{SPREADSHEET_CONVERSION_MARKER}\n"
    return f"{SPREADSHEET_CONVERSION_MARKER}\n\n{body}\n"


def xls_cell_text(value) -> str:
    if value is None or value == "":
        return ""
    return format_cell_value(value)
